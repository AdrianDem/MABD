import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import os
import time
import threading

def rmfile(filepath):
    """Remove file if it exists."""
    if os.path.isfile(filepath):
        os.remove(filepath)

def read_excel_file(file_path):
    """Reads data from an Excel file and returns it as a list of lists."""
    print(f"Попытка открыть файл: {file_path}")  # Отладочный вывод
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        return [[cell.value for cell in row] for row in sheet.iter_rows()]
    except FileNotFoundError:
        messagebox.showerror("Ошибка", f"Файл не найден: {file_path}")
        return None
    except openpyxl.utils.exceptions.InvalidFileException:
        messagebox.showerror("Ошибка", f"Неверный формат файла Excel: {file_path}")
        return None
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при открытии файла: {e}")
        return None

def convert_to_strings(values):
    """Converts all values in a list of lists to strings."""
    return [[str(value).strip() if value is not None else '' for value in row] for row in values]

def create_new_workbook_and_sheet():
    """Creates a new Excel workbook and sheet."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    return workbook, sheet

def adjust_column_width(sheet):
    """Adjusts the width of columns in a sheet based on the maximum length of the values."""
    for column in sheet.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value is not None), default=0)
        sheet.column_dimensions[column[0].column_letter].width = max_length + 2

def write_values_to_excel(values, output_file):
    """Writes values to an Excel file, adjusts column width, and saves the file."""
    try:
        workbook, sheet = create_new_workbook_and_sheet()
        for row in values:
            sheet.append(row)
        adjust_column_width(sheet)
        workbook.save(output_file)
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при записи в файл: {e}")

def find_unique_rows(values1, values2):
    """Finds unique rows between two sets of values."""
    set1 = set(map(tuple, values1))
    set2 = set(map(tuple, values2))
    return list(set1 - set2) + list(set2 - set1)

def choose_file():
    """Opens a file dialog to choose an Excel file."""
    try:
        file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])
        print(f"Выбран файл: {file_path}")  # Отладочный вывод
        return file_path
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка в choose_file: {e}")
        return None

def save_file_as():
    """Opens a file dialog to save an Excel file."""
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        print(f"Файл для сохранения: {file_path}")  # Отладочный вывод
        return file_path
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка в save_file_as: {e}")
        return None

def animate_loading_text():
    """Animates the loading text."""
    loading_text = loading_label.cget("text")
    if loading_text.endswith("..."):
        loading_text = "Загрузка"
    else:
        loading_text += "."
    loading_label.config(text=loading_text)
    loading_label.after(500, animate_loading_text)  # Планирование следующего обновления

def start_process():
    """Starts the process of finding differences between Excel files."""
    # Display loading message
    loading_label.config(text="Загрузка")
    animate_loading_text()

    # Get file paths from entry fields
    file1 = file1_entry.get()
    file2 = file2_entry.get()

    # Check if file paths are provided
    if not file1 or not file2:
        messagebox.showerror("Ошибка", "Пожалуйста, выберите оба файла.")
        stop_loading()
        return

    # Read data from Excel files
    data1 = read_excel_file(file1)
    if data1 is None:
        stop_loading()
        return

    data2 = read_excel_file(file2)
    if data2 is None:
        stop_loading()
        return

    # Convert data to strings
    values1 = convert_to_strings(data1)
    values2 = convert_to_strings(data2)

    # Find unique rows
    unique_values = find_unique_rows(values1, values2)

    # Save unique rows to a new Excel file
    output_file = save_file_as()
    if output_file:
        write_values_to_excel(unique_values, output_file)
        messagebox.showinfo("Успех", "Отчет успешно сохранен.")

    # Stop loading animation
    stop_loading()

def stop_loading():
    """Stops the loading animation."""
    loading_label.after_cancel(animate_loading_text)  # Отмена запланированных задач
    loading_label.config(text="")

def run_process():
    """Runs the process in a separate thread."""
    thread = threading.Thread(target=start_process)
    thread.start()

def close_window():
    """Closes the main window."""
    root.destroy()

# Main window setup
root = tk.Tk()
root.title("Сравнение Excel файлов")

# Title label
title_label = tk.Label(root, text="Программа сравнения двух Excel файлов")
title_label.grid(row=0, column=0, columnspan=3, padx=5, pady=5)

# First file selection
file1_label = tk.Label(root, text="Первый файл:")
file1_label.grid(row=1, column=0, padx=5, pady=5)
file1_entry = tk.Entry(root, width=50)
file1_entry.grid(row=1, column=1, padx=5, pady=5)
file1_button = tk.Button(root, text="Выбрать первый файл", command=lambda: file1_entry.insert(0, choose_file()))
file1_button.grid(row=1, column=2, padx=5, pady=5)

# Second file selection
file2_label = tk.Label(root, text="Второй файл:")
file2_label.grid(row=2, column=0, padx=5, pady=5)
file2_entry = tk.Entry(root, width=50)
file2_entry.grid(row=2, column=1, padx=5, pady=5)
file2_button = tk.Button(root, text="Выбрать второй файл", command=lambda: file2_entry.insert(0, choose_file()))
file2_button.grid(row=2, column=2, padx=5, pady=5)

# Buttons
compare_button = tk.Button(root, text="Найти различия", command=run_process)
compare_button.grid(row=3, column=1, padx=5, pady=5)
close_button = tk.Button(root, text="Закрыть", command=close_window)
close_button.grid(row=3, column=2, padx=5, pady=5)

# Loading label
loading_label = tk.Label(root, text="")
loading_label.grid(row=4, column=0, columnspan=3, padx=5, pady=5)

# Main loop
root.mainloop()
